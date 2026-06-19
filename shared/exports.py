from io import BytesIO
from django.http import HttpResponse


def _resolve(obj, accessor):
    """Resuelve un campo anidado ('brand.name') o un callable sobre el objeto."""
    if callable(accessor):
        return accessor(obj)
    value = obj
    for part in accessor.split('.'):
        value = getattr(value, part, '')
        if callable(value):
            value = value()
    return '' if value is None else value


def export_to_pdf(queryset, columns, title):
    """
    Exporta un queryset a PDF.

    columns: lista de tuplas (encabezado, accessor)
             accessor puede ser un string de campo ('brand.name') o un callable.
    Ajusta automáticamente el tamaño de fuente y ancho de columnas según la cantidad.
    """
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from datetime import datetime

    buffer = BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page_size,
        rightMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        title=title,
    )
    styles = getSampleStyleSheet()

    n = len(columns)
    # Tamaño de fuente adaptativo según cantidad de columnas
    font_h = max(7, 12 - max(0, n - 5))
    font_d = max(6, 10 - max(0, n - 5))

    data = [[col[0] for col in columns]]
    for obj in queryset:
        data.append([str(_resolve(obj, col[1])) for col in columns])

    usable_width = page_size[0] - 3 * cm
    col_width = usable_width / n

    table = Table(data, colWidths=[col_width] * n, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#343a40')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), font_h),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('FONTSIZE', (0, 1), (-1, -1), font_d),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])
    for i in range(1, len(data)):
        bg = colors.HexColor('#f8f9fa') if i % 2 == 0 else colors.white
        style.add('BACKGROUND', (0, i), (-1, i), bg)

    table.setStyle(style)

    generated = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    doc.build([
        Paragraph(title, styles['Title']),
        Paragraph(generated, styles['Normal']),
        Spacer(1, 0.4 * cm),
        table,
    ])

    buffer.seek(0)
    slug = title.lower().replace(' ', '_')
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{slug}.pdf"'
    return response


def export_to_excel(queryset, columns, title):
    """
    Exporta un queryset a Excel (.xlsx).

    columns: lista de tuplas (encabezado, accessor)
             accessor puede ser un string de campo ('brand.name') o un callable.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]

    header_fill = PatternFill(start_color='343A40', end_color='343A40', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    alt_fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')

    ws.append([col[0] for col in columns])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, obj in enumerate(queryset, start=2):
        ws.append([str(_resolve(obj, col[1])) for col in columns])
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = alt_fill

    for col_cells in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col_cells), default=10)
        ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 50)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    slug = title.lower().replace(' ', '_')
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{slug}.xlsx"'
    return response
